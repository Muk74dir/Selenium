from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
import time

wb = load_workbook('D:/Selenium/data.xlsx')
day = input("Enter Day: ")

try:
    file = wb[day]
    print(f"Sheet {day} found!")
except Exception as e:
    print(f"Enter Valid Day :--> {str(e)}")
    exit()

service = Service('C:/Program Files (x86)/chromedriver.exe')
driver = webdriver.Chrome(service=service, options=webdriver.ChromeOptions())

driver.get('https://google.com')

input_element = driver.find_element(By.CLASS_NAME, 'gLFyf')

for i in range(3, 13):
    keyword = file.cell(row=i, column=3).value
    input_element.send_keys(keyword)
    time.sleep(1)
    search_keywords = driver.find_elements(By.CLASS_NAME, 'aajZCb')
    search_values = [x.text for x in search_keywords]
    cleaned_values = search_values[0].split('\n')
    longest_value = max(cleaned_values, key=len)
    shortest_value = min(cleaned_values, key=len)
    file[f'D{i}'] = longest_value
    file[f'E{i}'] = shortest_value
    input_element.clear()

try:
    wb.save('D:/Selenium/data.xlsx')
    print("Data saved successfully!")
except Exception as e:
    print(f"An error occurred while saving the data: {str(e)}")

driver.quit()

